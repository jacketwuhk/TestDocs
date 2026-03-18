"""
ETC 通行记录分析脚本
分析起止站点相同但收费金额不同的记录，汇总为一行显示每个不同价格的记录数及与最低价的差额
"""

import csv
import os
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_FILL_COLOR = "00000000"
WHITE_FILL_COLOR = "FFFFFFFF"

INPUT_FILE = "ETC.csv"
OUTPUT_FILE = "ETC_analysis.xlsx"


def load_data(filename):
    """读取 ETC.csv，返回按起止站点分组的金额列表"""
    route_amounts = defaultdict(list)
    total_rows = 0
    with open(filename, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            route = row["起止站点"].strip()
            amount = float(row["金额"])
            route_amounts[route].append(amount)
            total_rows += 1
    return route_amounts, total_rows


def analyze(route_amounts):
    """找出有多个不同金额的起止站点，返回汇总数据列表"""
    results = []
    for route, amounts in sorted(route_amounts.items()):
        price_counter = Counter(amounts)
        if len(price_counter) <= 1:
            continue  # 只有一种价格，跳过

        sorted_prices = sorted(price_counter.keys())
        min_price = sorted_prices[0]
        total_count = sum(price_counter.values())

        entry = {
            "起止站点": route,
            "总记录数": total_count,
            "prices": [
                {
                    "价格": price,
                    "记录数": price_counter[price],
                    "与最低价差额": round(price - min_price, 2),
                }
                for price in sorted_prices
            ],
        }
        results.append(entry)
    return results


def write_excel(results, max_prices, output_file):
    """将分析结果写入 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ETC差异分析"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_main = PatternFill("solid", fgColor="366092")    # 深蓝色：主标题列
    header_fill_price = [
        PatternFill("solid", fgColor="70AD47"),  # 绿色
        PatternFill("solid", fgColor="ED7D31"),  # 橙色
        PatternFill("solid", fgColor="FFC000"),  # 金色
        PatternFill("solid", fgColor="FF0000"),  # 红色
        PatternFill("solid", fgColor="7030A0"),  # 紫色
    ]
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── 第1行：大标题 ──
    total_cols = 2 + max_prices * 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value="ETC 起止站点相同但收费金额不同记录汇总")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = center

    # ── 第2行：列标题（第一层） ──
    ws.cell(row=2, column=1, value="起止站点").alignment = center
    ws.cell(row=2, column=2, value="总记录数").alignment = center
    for i in range(max_prices):
        col_start = 3 + i * 3
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + 2)
        cell = ws.cell(row=2, column=col_start, value=f"价格 {i + 1}")
        cell.alignment = center

    # ── 第3行：列标题（第二层） ──
    ws.cell(row=3, column=1, value="起止站点").alignment = center
    ws.cell(row=3, column=2, value="总记录数").alignment = center
    for i in range(max_prices):
        col_start = 3 + i * 3
        ws.cell(row=3, column=col_start, value="金额（元）").alignment = center
        ws.cell(row=3, column=col_start + 1, value="记录数").alignment = center
        ws.cell(row=3, column=col_start + 2, value="与最低价差额（元）").alignment = center

    # 合并第2、3行中"起止站点"和"总记录数"
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

    # 为标题行设置样式
    for col in range(1, total_cols + 1):
        for row in (2, 3):
            cell = ws.cell(row=row, column=col)
            if col <= 2:
                cell.fill = header_fill_main
                cell.font = header_font
            else:
                price_idx = (col - 3) // 3
                cell.fill = header_fill_price[price_idx % len(header_fill_price)]
                cell.font = header_font
            cell.border = thin
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = center

    # ── 数据行 ──
    data_start_row = 4
    for r_idx, entry in enumerate(results):
        row = data_start_row + r_idx
        fill_row = PatternFill("solid", fgColor="DEEAF1") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        ws.cell(row=row, column=1, value=entry["起止站点"]).alignment = left
        ws.cell(row=row, column=2, value=entry["总记录数"]).alignment = center

        for i, p in enumerate(entry["prices"]):
            col_start = 3 + i * 3
            ws.cell(row=row, column=col_start, value=p["价格"]).alignment = center
            ws.cell(row=row, column=col_start + 1, value=p["记录数"]).alignment = center
            diff_cell = ws.cell(row=row, column=col_start + 2, value=p["与最低价差额"])
            diff_cell.alignment = center
            if p["与最低价差额"] > 0:
                diff_cell.font = Font(color="C00000", bold=True)

        # 样式：整行
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in (DEFAULT_FILL_COLOR, WHITE_FILL_COLOR):
                cell.fill = fill_row
            cell.border = thin

    # ── 列宽 ──
    ws.column_dimensions[get_column_letter(1)].width = 36  # 起止站点
    ws.column_dimensions[get_column_letter(2)].width = 10  # 总记录数
    for i in range(max_prices):
        col_start = 3 + i * 3
        ws.column_dimensions[get_column_letter(col_start)].width = 14      # 金额
        ws.column_dimensions[get_column_letter(col_start + 1)].width = 10  # 记录数
        ws.column_dimensions[get_column_letter(col_start + 2)].width = 18  # 差额

    # ── 冻结前3行 ──
    ws.freeze_panes = "A4"

    # ── 第1行高度 ──
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    wb.save(output_file)
    print(f"分析结果已保存到：{output_file}")
    print(f"共找到 {len(results)} 条起止站点相同但收费金额不同的记录组")


def main():
    print(f"正在读取 {INPUT_FILE} ...")
    route_amounts, total_rows = load_data(INPUT_FILE)
    print(f"共读取 {total_rows} 条记录，{len(route_amounts)} 个不同起止站点")

    print("正在分析差异数据...")
    results = analyze(route_amounts)

    # 找出最多价格种数（决定列数）
    max_prices = max(len(e["prices"]) for e in results) if results else 1

    print(f"找到 {len(results)} 个起止站点有差异价格，最多有 {max_prices} 种不同价格")
    print(f"正在生成 {OUTPUT_FILE} ...")
    write_excel(results, max_prices, OUTPUT_FILE)


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
